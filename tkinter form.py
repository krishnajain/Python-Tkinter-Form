from openpyxl import *
from tkinter import *

wb = load_workbook('YOUR EXCEL LOCATION.xlsx') 

sheet = wb.active

def excel():  
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 10
    sheet.column_dimensions['C'].width = 10
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 20
    sheet.column_dimensions['F'].width = 40
    sheet.column_dimensions['G'].width = 50
   
    sheet.cell(row=1, column=1).value = "Name"
    sheet.cell(row=1, column=2).value = "Class"
    sheet.cell(row=1, column=3).value = "Section"
    sheet.cell(row=1, column=4).value = "Roll No."
    sheet.cell(row=1, column=5).value = "Contact Number"
    sheet.cell(row=1, column=6).value = "Email id"
    sheet.cell(row=1, column=7).value = "Address"
    
def focus1(event):
    class_field.focus_set() 

def focus2(event):
    section_field.focus_set() 
   
def focus3(event):
    roll_no_field.focus_set() 
  
def focus4(event): 
    contact_no_field.focus_set() 
 
def focus5(event): 
    email_id_field.focus_set() 
   
def focus6(event): 
    address_field.focus_set() 
 
def clear():  
    name_field.delete(0, END) 
    class_field.delete(0, END) 
    section_field.delete(0, END) 
    roll_no_field.delete(0, END) 
    contact_no_field.delete(0, END) 
    email_id_field.delete(0, END) 
    address_field.delete(0, END)
    
def insert(): 

    if (name_field.get() == "" and
        class_field.get() == "" and
        section_field.get() == "" and
        roll_no_field.get() == "" and
        contact_no_field.get() == "" and
        email_id_field.get() == "" and
        address_field.get() == ""): 
              
        print("empty input") 
  
    else:  
        current_row = sheet.max_row 
        current_column = sheet.max_column  
        sheet.cell(row=current_row + 1, column=1).value = name_field.get() 
        sheet.cell(row=current_row + 1, column=2).value = class_field.get() 
        sheet.cell(row=current_row + 1, column=3).value = section_field.get() 
        sheet.cell(row=current_row + 1, column=4).value = roll_no_field.get() 
        sheet.cell(row=current_row + 1, column=5).value = contact_no_field.get() 
        sheet.cell(row=current_row + 1, column=6).value = email_id_field.get() 
        sheet.cell(row=current_row + 1, column=7).value = address_field.get()
        
        wb.save('YOUR EXCEL LOCATION.xlsx') 
        name_field.focus_set()  
        clear() 
 
if __name__ == "__main__": 
    root = Tk() 
    root.configure(background='green')  
    root.title("Student Registration Form")  
    root.geometry("500x400") 
    excel() 
    heading = Label(root, text="Form", bg="light green") 
    name = Label(root, text="Name", bg="light green") 
    Class = Label(root, text="Class", bg="light green") 
    section = Label(root, text="Section", bg="light green") 
    roll_no = Label(root, text="Roll No.", bg="light green")  
    contact_no = Label(root, text="Contact No.", bg="light green")  
    email_id = Label(root, text="Email id", bg="light green")  
    address = Label(root, text="Address", bg="light green") 
    heading.grid(row=0, column=1) 
    name.grid(row=1, column=0) 
    Class.grid(row=2, column=0) 
    section.grid(row=3, column=0) 
    roll_no.grid(row=4, column=0) 
    contact_no.grid(row=5, column=0) 
    email_id.grid(row=6, column=0) 
    address.grid(row=7, column=0) 

    name_field = Entry(root) 
    class_field = Entry(root) 
    section_field = Entry(root) 
    roll_no_field = Entry(root) 
    contact_no_field = Entry(root) 
    email_id_field = Entry(root) 
    address_field = Entry(root) 

    name_field.bind("<Return>", focus1) 
    class_field.bind("<Return>", focus2) 
    section_field.bind("<Return>", focus3)  
    roll_no_field.bind("<Return>", focus4)  
    contact_no_field.bind("<Return>", focus5) 
    email_id_field.bind("<Return>", focus6) 
 
    name_field.grid(row=1, column=1, ipadx="100") 
    class_field.grid(row=2, column=1, ipadx="100") 
    section_field.grid(row=3, column=1, ipadx="100") 
    roll_no_field.grid(row=4, column=1, ipadx="100") 
    contact_no_field.grid(row=5, column=1, ipadx="100") 
    email_id_field.grid(row=6, column=1, ipadx="100") 
    address_field.grid(row=7, column=1, ipadx="100") 

    excel() 
    submit = Button(root, text="Submit", fg="Black", 
                            bg="Red", command=insert) 
    submit.grid(row=8, column=1) 
    root.mainloop() 
